import argparse
import datetime
import os

import pandas as pd
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt
from scipy import stats

parser = argparse.ArgumentParser()
parser.add_argument('-fn', '--filename',
                    type=str,
                    default=f"../data/correlation_quotes_{datetime.datetime.now().strftime('%B')}.xlsx",
                    help='pass the excel file where the quotes are stored')

args = parser.parse_args()


# TODO: Create a new script to read the data from quotes file not correlation file and stores in temp image and stores in sheet (4 chart max on 1 row) (display_chart.py)
def draw_chart(sheet_name: str, corr_array: pd.DataFrame, book_writer):
    chart_sheet_title = f"Correlation_chart_{sheet_name}"
    book_writer.book.create_sheet(title=chart_sheet_title)
    chart_worksheet = book_writer.book[chart_sheet_title]

    col_count = corr_array.shape[1]
    row_start = 1
    col_start = 1
    chart_width = 20
    chart_height = 7
    charts_per_row = 4  # Number of charts per row
    chart_index = 0  # Initialize chart index

    for i in range(col_count):
        for j in range(i + 1, col_count):
            if i != j:  # skip self-correlation
                currency1 = corr_array.columns[i]
                currency2 = corr_array.columns[j]

                # Create a scatter plot using Matplotlib
                plt.figure(figsize=(chart_width, chart_height))
                plt.scatter(corr_array[currency1], corr_array[currency2])
                plt.title(f'Scatter Plot: {currency1} vs {currency2}')
                plt.xlabel(currency1)
                plt.ylabel(currency2)

                # Save the plot to a temporary file
                temp_file = f'../data/temp_charts/temp_plot_{currency1}_{currency2}.png'
                plt.savefig(temp_file)
                plt.close()

                # Load the saved image into Excel worksheet
                img = Image(f"{temp_file}")
                chart_worksheet.add_image(img, f'A{row_start + (chart_index // charts_per_row) * (chart_height + 2)}')

                # Increment chart index
                chart_index += 1
                # print(f"Chart{currency1} vs {currency2} saved for {sheet_name} TimeFrame.")
    [os.remove(f'../data/temp_charts/{f}') for f in os.listdir('../data/temp_charts/')]
    print(f"Charts stored in {chart_sheet_title}\n\n")


def corr_calculation(fname: str = args.filename):
    excel_path = fname
    sheets = pd.read_excel(excel_path, sheet_name=None, engine='openpyxl')

    # where the calculated coefficient will be stored
    correlation_file = f"../data/correlation_coefficient_{fname.split('_')[-1]}"

    # Create an Excel writer object to write multiple sheets
    with pd.ExcelWriter(correlation_file, engine='openpyxl') as writer:
        # Iterate through each sheet
        for sheet_name, df in sheets.items():
            df_numpy = df.to_numpy()

            # Create an empty DataFrame to store the correlation matrix
            column_names = df.columns.tolist()
            correlation_matrix = pd.DataFrame(index=column_names, columns=column_names)

            column_count = df_numpy.shape[1]
            for i in range(column_count):
                for j in range(column_count):
                    if i == j:  # self-correlation is always 1
                        correlation_matrix.iloc[i, i] = 1.0
                    else:
                        r, p = stats.pearsonr(df_numpy[:, i], df_numpy[:, j])
                        correlation_matrix.iloc[i, j] = round(r * 100, 3)

            correlation_matrix.to_excel(writer, sheet_name=f'Correlation_{sheet_name}', index=True)
            print(f"Correlation Coefficient for sheet '{sheet_name}' stored in the workbook.")

            # Get the workbook and the last written sheet
            workbook = writer.book
            worksheet = workbook[f'Correlation_{sheet_name}']

            # Define fill styles
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            light_green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
            light_light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            light_red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            light_light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Apply background colors based on correlation values
            for row in range(2, len(column_names) + 2):  # Excel rows start at 1, header row is row 1
                for col in range(2, len(column_names) + 2):  # Column indices for values
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        if 70 <= cell_value <= 100:
                            worksheet.cell(row=row, column=col).fill = green_fill
                        elif 30 <= cell_value < 70:
                            worksheet.cell(row=row, column=col).fill = light_green_fill
                        elif 0 <= cell_value < 30:
                            worksheet.cell(row=row, column=col).fill = light_light_green_fill
                        elif -100 <= cell_value <= -70:
                            worksheet.cell(row=row, column=col).fill = red_fill
                        elif -70 < cell_value <= -30:
                            worksheet.cell(row=row, column=col).fill = light_red_fill
                        elif -29 <= cell_value < -10:
                            worksheet.cell(row=row, column=col).fill = light_light_red_fill
                        elif -10 <= cell_value <= 10:
                            worksheet.cell(row=row, column=col).fill = gray_fill
            print(f"Values formatted and stored in sheet 'Correlation_{sheet_name}' stored at {correlation_file}.")
            draw_chart(sheet_name, correlation_matrix, writer)

    print(f"All Correlation Coefficient stored in {correlation_file}")


corr_calculation()
