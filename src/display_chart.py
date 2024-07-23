import os
import shutil

import pandas as pd
import datetime
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from pandas.plotting import register_matplotlib_converters
from openpyxl.drawing.image import Image
import argparse

register_matplotlib_converters()

parser = argparse.ArgumentParser()
parser.add_argument('-fn', '--filename',
                    type=str,
                    default=f"../data/correlation_quotes_{datetime.datetime.now().strftime('%B_%Y-%m-%d_%H-%M')}.xlsx",
                    help='pass the excel file where the quotes are stored')

args = parser.parse_args()


# TODO: Create a new script to read the data from quotes file not correlation file and stores in temp image and stores in sheet (4 chart max on 1 row) (display_chart.py)
def draw_chart(correlation_file: str = args.filename):
    # Create an Excel writer object to write multiple sheets

    excel_file = pd.ExcelFile(correlation_file)
    sheetnames = excel_file.sheet_names

    # Create a new Excel workbook
    wb = load_workbook(correlation_file)

    # Iterate through each sheet
    for sheetname in sheetnames:
        # Read each sheet into a DataFrame
        df = pd.read_excel(excel_file, sheet_name=sheetname, header=0)
        ws = wb.create_sheet(title=f'{sheetname}_Chart')

        # draw chart for each currency pair
        for ticker in df.columns:
            if ticker != df.index:
                plt.figure(figsize=(10, 6))
                plt.scatter(df.index, df[ticker], marker='o', label=ticker)
                plt.title(f'{ticker} Quotes')
                plt.xlabel('Index')
                plt.ylabel('Quote')
                plt.legend()
                plt.grid(True)
                plt.tight_layout()

                # Save plot as image
                img_path = f'../data/{ticker}_plot.png'
                plt.savefig(img_path)

                # Insert image into Excel sheet
                img = Image(img_path)
                ws.add_image(img, f'A1')  # Insert image at cell A1

                # Close plot figure and remove image file
                plt.close()
                os.remove(img_path)
    wb.save(correlation_file)

    # with pd.ExcelWriter(correlation_file, engine='openpyxl') as writer:
    #     chart_sheet_title = f""
    #     book_writer.book.create_sheet(title=chart_sheet_title)
    #     chart_worksheet = book_writer.book[chart_sheet_title]
    #
    #     col_count = corr_array.shape[1]
    #     row_start = 1
    #     col_start = 1
    #     chart_width = 20
    #     chart_height = 7
    #     charts_per_row = 4  # Number of charts per row
    #     chart_index = 0  # Initialize chart index
    #
    #     if os.path.exists('../data/temp_charts'):
    #         shutil.rmtree('../data/temp_charts')
    #     else:
    #         os.mkdir('../data/temp_charts')
    #
    #     for i in range(col_count):
    #         for j in range(i + 1, col_count):
    #             if i != j:  # skip self-correlation
    #                 currency1 = corr_array.columns[i]
    #                 currency2 = corr_array.columns[j]
    #
    #                 # Create a scatter plot using Matplotlib
    #                 plt.figure(figsize=(chart_width, chart_height))
    #                 plt.scatter(corr_array[currency1], corr_array[currency2])
    #                 plt.title(f'Scatter Plot: {currency1} vs {currency2}')
    #                 plt.xlabel(currency1)
    #                 plt.ylabel(currency2)
    #
    #                 # Save the plot to a temporary file
    #                 temp_file = f'../data/temp_charts/temp_plot_{currency1}_{currency2}.png'
    #                 plt.savefig(temp_file)
    #                 plt.close()
    #
    #                 # Load the saved image into Excel worksheet
    #                 img = Image(f"{temp_file}")
    #                 chart_worksheet.add_image(img, f'A{row_start + (chart_index // charts_per_row) * (chart_height + 2)}')
    #
    #                 # Increment chart index
    #                 chart_index += 1
    #                 # print(f"Chart{currency1} vs {currency2} saved for {sheet_name} TimeFrame.")


draw_chart()
